#!/usr/bin/env python3
"""
JSON Advertisement Information Extractor - UPDATED VERSION
- Supports Gemini API (google-genai), OpenAI API, and Hugging Face local models
- Extracts structured information from advertisement images using JSON prompting
- FIXED: Properly handles Gemini's markdown-wrapped JSON responses
- Outputs results as JSON and a comprehensive Excel file
"""

import os
import json
import argparse
import shutil
from pathlib import Path
from PIL import Image
import time
import logging
from datetime import datetime
import pandas as pd
import base64
import io
import torch

# Try importing all providers
try:
    from google import genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import torch
    from transformers import (
        LlavaNextProcessor, LlavaNextForConditionalGeneration,
        LlavaForConditionalGeneration, LlavaProcessor,
        AutoProcessor, AutoModelForCausalLM,
        Gemma3nForConditionalGeneration
    )
    HF_AVAILABLE = True
except ImportError:
    HF_AVAILABLE = False

# Simple .env loader
def load_env_file():
    try:
        with open('.env', 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    value = value.strip().strip('"').strip("'")
                    os.environ[key.strip()] = value
    except FileNotFoundError:
        pass

load_env_file()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def move_to_device(obj, device):
    """Improved device movement function"""
    if torch.is_tensor(obj):
        return obj.to(device)
    elif isinstance(obj, dict):
        return {k: move_to_device(v, device) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [move_to_device(v, device) for v in obj]
    elif isinstance(obj, tuple):
        return tuple(move_to_device(v, device) for v in obj)
    else:
        return obj

class BaseVisionAnalyzer:
    """Base class for all vision analyzers"""
    
    def __init__(self, model_name):
        self.model_name = model_name
        self.api_calls_count = 0
        # JSON prompting for advertisement analysis
        self.json_extraction_prompt = """
Analyze this advertisement image and extract key information in the following JSON format. 
Be precise and only include information that is clearly visible or strongly implied in the image.

Return ONLY a valid JSON object with this exact structure:

{
    "brand": {
        "name": "Brand name if visible",
        "logo_present": true/false,
        "logo_position": "position of logo if present"
    },
    "product": {
        "category": "Product category (e.g., car, phone, clothing)",
        "name": "Specific product name if shown",
        "model": "Product model if specified",
        "features_highlighted": ["list of key features shown"]
    },
    "visual_elements": {
        "dominant_colors": ["primary colors in the ad"],
        "layout_style": "Description of visual layout",
        "image_style": "photographic/illustrated/minimal/etc",
        "people_present": true/false,
        "lifestyle_context": "Context shown (home, office, outdoors, etc)"
    },
    "text_content": {
        "headline": "Main headline text",
        "tagline": "Brand tagline or slogan",
        "call_to_action": "CTA text like 'Buy Now', 'Learn More'",
        "price_information": "Any pricing shown",
        "other_text": ["Other significant text elements"]
    },
    "marketing_approach": {
        "target_audience": "Who this ad targets",
        "emotional_appeal": "Fear/Joy/Status/Convenience/etc",
        "value_proposition": "Main benefit emphasized",
        "advertising_style": "luxury/budget/tech-focused/lifestyle/etc"
    },
    "composition": {
        "focal_point": "What draws attention first",
        "text_to_image_ratio": "text-heavy/balanced/image-heavy",
        "layout_type": "grid/centered/asymmetric/etc"
    }
}

Respond with ONLY the JSON object. Do not include any other text, explanations, or markdown formatting.
"""
    
    def analyze_image_json_extraction(self, image_path):
        """Run the JSON extraction prompt on the image - to be implemented by subclasses"""
        raise NotImplementedError
    
    def _get_image_dimensions(self, image_path):
        try:
            with Image.open(image_path) as img:
                return {"width": img.width, "height": img.height}
        except:
            return {"width": 0, "height": 0}

class GeminiAnalyzer(BaseVisionAnalyzer):
    def __init__(self, model_name="gemini-2.0-flash-001"):
        super().__init__(model_name)
        self._initialize_model()
    
    def _initialize_model(self):
        if not GEMINI_AVAILABLE:
            raise ImportError("google-genai not installed. Run: pip install google-genai")
        api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY or GOOGLE_API_KEY not found in environment")
        self.client = genai.Client(api_key=api_key)
        logger.info(f"✅ Gemini model initialized: {self.model_name}")
    
    def analyze_image_json_extraction(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        logger.info(f"Analyzing: {image_name} for JSON extraction (Gemini)")
        try:
            with open(image_path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')
            image_format = Path(image_path).suffix.lower().replace('.', '')
            if image_format == 'jpg':
                image_format = 'jpeg'
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=[
                    self.json_extraction_prompt,
                    {"inline_data": {"mime_type": f"image/{image_format}", "data": image_data}}
                ]
            )
            json_response = response.text.strip()
            
            # Debug: Show first 200 chars of response
            logger.info(f"Raw response preview: {repr(json_response[:200])}")
            logger.info(f"Response length: {len(json_response)}")
            
            # Strip markdown code blocks if present
            clean_json = json_response
            if clean_json.startswith('```json'):
                clean_json = clean_json[7:]  # Remove ```json
            if clean_json.startswith('```'):
                clean_json = clean_json[3:]   # Remove ```
            if clean_json.endswith('```'):
                clean_json = clean_json[:-3]  # Remove trailing ```
            clean_json = clean_json.strip()
            
            # Try to parse JSON to validate
            try:
                if not clean_json:
                    raise json.JSONDecodeError("Empty response after cleaning", "", 0)
                parsed_json = json.loads(clean_json)
                structured_data = parsed_json
                parsing_success = True
                logger.info("✅ Successfully parsed JSON response")
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse JSON response: {e}")
                logger.warning(f"Cleaned JSON: {repr(clean_json[:200])}")
                structured_data = {"parsing_error": str(e), "raw_response": json_response, "cleaned_json": clean_json}
                parsing_success = False
            
            self.api_calls_count += 1
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'gemini',
                'extracted_data': structured_data,
                'raw_json_response': json_response,
                'parsing_success': parsing_success,
                'analysis_timestamp': datetime.now().isoformat(),
                'processing_metadata': {
                    'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                    'image_dimensions': self._get_image_dimensions(image_path)
                }
            }
        except Exception as e:
            logger.error(f"Error analyzing JSON extraction: {e}")
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'gemini',
                'error': str(e),
                'analysis_timestamp': datetime.now().isoformat()
            }

class OpenAIAnalyzer(BaseVisionAnalyzer):
    def __init__(self, model_name="gpt-4o"):
        super().__init__(model_name)
        self._initialize_model()
    
    def _initialize_model(self):
        if not OPENAI_AVAILABLE:
            raise ImportError("openai not installed. Run: pip install openai")
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not found in environment")
        self.client = openai.OpenAI(api_key=api_key)
        logger.info(f"✅ OpenAI model initialized: {self.model_name}")
    
    def _encode_image_base64(self, image_path):
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    
    def analyze_image_json_extraction(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        logger.info(f"Analyzing: {image_name} for JSON extraction (OpenAI)")
        try:
            base64_image = self._encode_image_base64(image_path)
            response = self.client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self.json_extraction_prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                        ]
                    }
                ],
                max_tokens=1500
            )
            json_response = response.choices[0].message.content.strip()
            
            # Debug: Show first 200 chars of response
            logger.info(f"Raw response preview: {repr(json_response[:200])}")
            logger.info(f"Response length: {len(json_response)}")
            
            # Strip markdown code blocks if present
            clean_json = json_response
            if clean_json.startswith('```json'):
                clean_json = clean_json[7:]  # Remove ```json
            if clean_json.startswith('```'):
                clean_json = clean_json[3:]   # Remove ```
            if clean_json.endswith('```'):
                clean_json = clean_json[:-3]  # Remove trailing ```
            clean_json = clean_json.strip()
            
            # Try to parse JSON to validate
            try:
                if not clean_json:
                    raise json.JSONDecodeError("Empty response after cleaning", "", 0)
                parsed_json = json.loads(clean_json)
                structured_data = parsed_json
                parsing_success = True
                logger.info("✅ Successfully parsed JSON response")
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse JSON response: {e}")
                logger.warning(f"Cleaned JSON: {repr(clean_json[:200])}")
                structured_data = {"parsing_error": str(e), "raw_response": json_response, "cleaned_json": clean_json}
                parsing_success = False
            
            self.api_calls_count += 1
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'openai',
                'extracted_data': structured_data,
                'raw_json_response': json_response,
                'parsing_success': parsing_success,
                'analysis_timestamp': datetime.now().isoformat(),
                'processing_metadata': {
                    'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                    'image_dimensions': self._get_image_dimensions(image_path)
                }
            }
        except Exception as e:
            logger.error(f"Error analyzing JSON extraction: {e}")
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'openai',
                'error': str(e),
                'analysis_timestamp': datetime.now().isoformat()
            }

class HuggingFaceAnalyzer(BaseVisionAnalyzer):
    def __init__(self, model_name="llava-hf/llava-1.5-13b-hf", device_arg=None):
        super().__init__(model_name)
        # If device_arg is set, use it; else, use device_map='auto' for large models
        self.device_arg = device_arg
        if device_arg is not None:
            self.device = device_arg
            self.device_map = None
        else:
            self.device = None  # Will not move tensors manually
            self.device_map = "auto"
        self.model = None
        self.processor = None
        self._initialize_model()
    
    def _initialize_model(self):
        if not HF_AVAILABLE:
            raise ImportError("transformers/torch not installed. Run: pip install transformers torch")
        print(f"🔄 Loading model: {self.model_name}")
        if self.device_arg is not None:
            print(f"📍 Device: {self.device_arg}")
        else:
            print(f"📍 Device: device_map='auto' (multi-GPU or large model)")
        cache_dir = os.path.expanduser("~/.cache/huggingface/hub")
        model_cache_name = f"models--{self.model_name.replace('/', '--')}"
        model_cache_path = os.path.join(cache_dir, model_cache_name)
        if os.path.exists(model_cache_path):
            print(f"💾 Found cached model at: {model_cache_path}")
        else:
            print(f"⬇️ Model not cached - will download to: {model_cache_path}")
        
        # Load model with proper configuration including Gemma 3n
        if "llava-v1.6" in self.model_name or "llava-next" in self.model_name:
            self.processor = LlavaNextProcessor.from_pretrained(self.model_name, use_fast=False)
            self.model = LlavaNextForConditionalGeneration.from_pretrained(
                self.model_name,
                torch_dtype=torch.float16,
                device_map=self.device_map,
                low_cpu_mem_usage=True
            )
        elif "llava" in self.model_name:
            self.processor = LlavaProcessor.from_pretrained(self.model_name, use_fast=False)
            self.model = LlavaForConditionalGeneration.from_pretrained(
                self.model_name,
                torch_dtype=torch.float16,
                device_map=self.device_map,
                low_cpu_mem_usage=True
            )
        elif "gemma-3n" in self.model_name.lower():
            # Special handling for Gemma 3n vision models
            self.processor = AutoProcessor.from_pretrained(self.model_name, use_fast=False)
            self.model = Gemma3nForConditionalGeneration.from_pretrained(
                self.model_name,
                torch_dtype=torch.float16,
                device_map=self.device_map,
                low_cpu_mem_usage=True
            )
        else:
            self.processor = AutoProcessor.from_pretrained(self.model_name, use_fast=False)
            self.model = AutoModelForCausalLM.from_pretrained(
                self.model_name,
                torch_dtype=torch.float16,
                device_map=self.device_map,
                low_cpu_mem_usage=True
            )
        
        # Move model to device if specified
        if self.device_arg is not None:
            self.model.to(self.device_arg)
            
        logger.info(f"✅ Hugging Face model loaded successfully")
    
    def analyze_image_json_extraction(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        logger.info(f"Analyzing: {image_name} for JSON extraction (HuggingFace)")
        try:
            image = Image.open(image_path).convert('RGB')
            json_response = self._generate_json_extraction(image, self.json_extraction_prompt)
            
            # Debug: Show first 200 chars of response
            logger.info(f"Raw response preview: {repr(json_response[:200])}")
            logger.info(f"Response length: {len(json_response)}")
            
            # Strip markdown code blocks if present
            clean_json = json_response
            if clean_json.startswith('```json'):
                clean_json = clean_json[7:]  # Remove ```json
            if clean_json.startswith('```'):
                clean_json = clean_json[3:]   # Remove ```
            if clean_json.endswith('```'):
                clean_json = clean_json[:-3]  # Remove trailing ```
            clean_json = clean_json.strip()
            
            # Try to parse JSON to validate
            try:
                if not clean_json:
                    raise json.JSONDecodeError("Empty response after cleaning", "", 0)
                parsed_json = json.loads(clean_json)
                structured_data = parsed_json
                parsing_success = True
                logger.info("✅ Successfully parsed JSON response")
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse JSON response: {e}")
                logger.warning(f"Cleaned JSON: {repr(clean_json[:200])}")
                structured_data = {"parsing_error": str(e), "raw_response": json_response, "cleaned_json": clean_json}
                parsing_success = False
            
            self.api_calls_count += 1
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'huggingface',
                'extracted_data': structured_data,
                'raw_json_response': json_response,
                'parsing_success': parsing_success,
                'analysis_timestamp': datetime.now().isoformat(),
                'processing_metadata': {
                    'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                    'image_dimensions': self._get_image_dimensions(image_path)
                }
            }
        except Exception as e:
            logger.error(f"Error analyzing JSON extraction: {e}")
            return {
                'image_path': str(image_path),
                'image_name': image_name,
                'image_id': image_id,
                'model_used': self.model_name,
                'provider': 'huggingface',
                'error': str(e),
                'analysis_timestamp': datetime.now().isoformat()
            }
    
    def _generate_json_extraction(self, image, prompt):
        try:
            # Special handling for Gemma 3n vision models
            if "gemma-3n" in self.model_name.lower():
                # Try different image token formats for Gemma 3n
                possible_tokens = ["<start_of_image>", "<image_soft_token>"]
                
                for image_token in possible_tokens:
                    try:
                        # Use the specific image token format for Gemma 3n
                        formatted_prompt = f"{image_token} {prompt}"
                        
                        # Process the image and text together using direct processing
                        inputs = self.processor(
                            text=formatted_prompt,
                            images=image,
                            return_tensors="pt"
                        )
                        
                        # Check if we got the right number of image tokens
                        logger.info(f"Trying {image_token} format...")
                        logger.info(f"Input shape: {inputs['input_ids'].shape}")
                        
                        # Move inputs to device
                        if self.device_arg is not None:
                            for key in inputs:
                                if torch.is_tensor(inputs[key]):
                                    inputs[key] = inputs[key].to(self.device_arg)
                        else:
                            model_device = next(self.model.parameters()).device
                            for key in inputs:
                                if torch.is_tensor(inputs[key]):
                                    inputs[key] = inputs[key].to(model_device)
                        
                        # Generate response
                        input_len = inputs["input_ids"].shape[-1]
                        with torch.no_grad():
                            outputs = self.model.generate(
                                **inputs,
                                max_new_tokens=800,
                                do_sample=True,
                                temperature=0.3,
                                top_p=0.9,
                                pad_token_id=self.processor.tokenizer.eos_token_id
                            )
                        
                        # Decode only the new tokens
                        new_tokens = outputs[0][input_len:]
                        response = self.processor.decode(new_tokens, skip_special_tokens=True)
                        logger.info(f"Success with {image_token} format!")
                        return response.strip()
                        
                    except Exception as e:
                        logger.warning(f"Failed with {image_token}: {e}")
                        continue
                
                # If all formats failed, raise the last error
                raise Exception("All image token formats failed for Gemma 3n")
            
            elif "llava" in self.model_name.lower():
                # For LLaVA models, we need to include the <image> token in the prompt
                if "llava-v1.6" in self.model_name or "llava-next" in self.model_name:
                    # For LLaVA-NeXT models
                    formatted_prompt = f"<image>\nUSER: {prompt}\nASSISTANT:"
                else:
                    # For regular LLaVA models  
                    formatted_prompt = f"USER: <image>\n{prompt}\nASSISTANT:"
                
                # Process the image and text together
                inputs = self.processor(
                    text=formatted_prompt,
                    images=image,
                    return_tensors="pt"
                )
                
                # Move ALL inputs to the same device as the model
                if self.device_arg is not None:
                    # Move all tensors to the specified device
                    for key in inputs:
                        if torch.is_tensor(inputs[key]):
                            inputs[key] = inputs[key].to(self.device_arg)
                        elif isinstance(inputs[key], list):
                            inputs[key] = [t.to(self.device_arg) if torch.is_tensor(t) else t for t in inputs[key]]
                else:
                    # If using device_map="auto", ensure all inputs are on the same device as the model
                    model_device = next(self.model.parameters()).device
                    for key in inputs:
                        if torch.is_tensor(inputs[key]):
                            inputs[key] = inputs[key].to(model_device)
                        elif isinstance(inputs[key], list):
                            inputs[key] = [t.to(model_device) if torch.is_tensor(t) else t for t in inputs[key]]
                
                # Generate with proper parameters for JSON
                with torch.no_grad():
                    outputs = self.model.generate(
                        **inputs,
                        max_new_tokens=800,
                        do_sample=True,
                        temperature=0.3,
                        top_p=0.9,
                        pad_token_id=self.processor.tokenizer.eos_token_id
                    )
                
                # Decode response
                response = self.processor.decode(outputs[0], skip_special_tokens=True)
                
                # Clean up the response - remove the input prompt
                if "ASSISTANT:" in response:
                    response = response.split("ASSISTANT:")[-1].strip()
                elif formatted_prompt in response:
                    response = response.replace(formatted_prompt, "").strip()
                
                return response
            else:
                # For other models, use standard processing
                inputs = self.processor(images=image, text=prompt, return_tensors="pt")
                
                # Move ALL inputs to the correct device
                if self.device_arg is not None:
                    for key in inputs:
                        if torch.is_tensor(inputs[key]):
                            inputs[key] = inputs[key].to(self.device_arg)
                        elif isinstance(inputs[key], list):
                            inputs[key] = [t.to(self.device_arg) if torch.is_tensor(t) else t for t in inputs[key]]
                else:
                    # If using device_map="auto", ensure all inputs are on the same device as the model
                    model_device = next(self.model.parameters()).device
                    for key in inputs:
                        if torch.is_tensor(inputs[key]):
                            inputs[key] = inputs[key].to(model_device)
                        elif isinstance(inputs[key], list):
                            inputs[key] = [t.to(model_device) if torch.is_tensor(t) else t for t in inputs[key]]
                
                with torch.no_grad():
                    outputs = self.model.generate(
                        **inputs,
                        max_new_tokens=800,
                        do_sample=True,
                        temperature=0.3,
                        top_p=0.9,
                        pad_token_id=self.processor.tokenizer.eos_token_id
                    )
                
                response = self.processor.decode(outputs[0], skip_special_tokens=True)
                
                # Clean up the response
                if prompt in response:
                    response = response.replace(prompt, "").strip()
                
                return response
                
        except Exception as e:
            logger.error(f"Error in _generate_json_extraction: {e}")
            return f"Error generating response: {str(e)}"
            
def get_analyzer(model_name, provider=None, device_arg=None):
    if provider == 'hf':
        if not HF_AVAILABLE:
            raise ImportError("Hugging Face model requested but transformers/torch not installed")
        return HuggingFaceAnalyzer(model_name, device_arg=device_arg)
    elif provider == 'openai':
        if not OPENAI_AVAILABLE:
            raise ImportError("OpenAI model requested but openai library not installed")
        return OpenAIAnalyzer(model_name)
    elif provider == 'gemini':
        if not GEMINI_AVAILABLE:
            raise ImportError("Gemini model requested but google-genai not installed")
        return GeminiAnalyzer(model_name)
    # Fallback to old logic if provider is None
    if any(hf_indicator in model_name.lower() for hf_indicator in 
           ['llava', 'qwen', 'git-', 'blip', 'instructblip', 'fuyu', 'glm', 'thudm', 'gemma-3n']):
        if not HF_AVAILABLE:
            raise ImportError("Hugging Face model requested but transformers/torch not installed")
        return HuggingFaceAnalyzer(model_name, device_arg=device_arg)
    elif any(openai_indicator in model_name.lower() for openai_indicator in 
             ['gpt-4o', 'gpt-4-vision', 'gpt-4-turbo']):
        if not OPENAI_AVAILABLE:
            raise ImportError("OpenAI model requested but openai library not installed")
        return OpenAIAnalyzer(model_name)
    elif any(gemini_indicator in model_name.lower() for gemini_indicator in 
             ['gemini', 'flash', 'pro']):
        if not GEMINI_AVAILABLE:
            raise ImportError("Gemini model requested but google-genai not installed")
        return GeminiAnalyzer(model_name)
    else:
        if not GEMINI_AVAILABLE:
            raise ImportError("Default Gemini model requested but google-genai not installed")
        return GeminiAnalyzer(model_name)

def create_output_structure(dataset_dir, model_name, topic=None):
    if topic is None:
        dataset_path = Path(dataset_dir)
        if "cars" in str(dataset_path):
            topic = "cars"
        elif "electronics" in str(dataset_path):
            topic = "electronics"
        elif "fashion" in str(dataset_path):
            topic = "fashion"
        else:
            parts = dataset_path.parts
            if len(parts) >= 2:
                topic = parts[-2]
            else:
                topic = "general"
    clean_model_name = model_name.replace("/", "_").replace(":", "_").replace("-", "_")
    output_dir = Path("output") / "json_extraction" / clean_model_name / topic
    output_dir.mkdir(parents=True, exist_ok=True)
    return {
        'output_dir': output_dir,
        'topic': topic,
        'model_name': clean_model_name,
        'base_output': Path("output")
    }

def save_results_to_folder(all_results, output_structure):
    saved_files = []
    output_dir = output_structure['output_dir']
    for result in all_results:
        if 'error' not in result:
            filename = f"{result['image_id']}.json"
            file_path = output_dir / filename
            with open(file_path, 'w') as f:
                json.dump(result, f, indent=2)
            saved_files.append(file_path)
    return saved_files

def flatten_json_for_excel(extracted_data):
    """Flatten nested JSON for Excel display with sanitization"""
    def sanitize_text(text):
        """Remove problematic characters for Excel"""
        if isinstance(text, str):
            # Remove or replace characters that might cause issues
            return text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        return text
    
    if isinstance(extracted_data, dict):
        if 'parsing_error' in extracted_data:
            # If parsing failed, still try to show useful info
            return {
                "extraction_status": "parsing_failed", 
                "error": sanitize_text(extracted_data.get('parsing_error', '')),
                "raw_response_preview": sanitize_text(extracted_data.get('raw_response', '')[:500] + "..." if len(extracted_data.get('raw_response', '')) > 500 else extracted_data.get('raw_response', '')),
                "cleaned_json_preview": sanitize_text(extracted_data.get('cleaned_json', '')[:500] + "..." if len(extracted_data.get('cleaned_json', '')) > 500 else extracted_data.get('cleaned_json', ''))
            }
        else:
            # Successfully parsed JSON
            flattened = {"extraction_status": "success"}
            for key, value in extracted_data.items():
                if isinstance(value, dict):
                    for subkey, subvalue in value.items():
                        if isinstance(subvalue, list):
                            flattened[f"{key}_{subkey}"] = sanitize_text("; ".join(map(str, subvalue)))
                        else:
                            flattened[f"{key}_{subkey}"] = sanitize_text(str(subvalue) if subvalue is not None else "")
                elif isinstance(value, list):
                    flattened[key] = sanitize_text("; ".join(map(str, value)))
                else:
                    flattened[key] = sanitize_text(str(value) if value is not None else "")
            return flattened
    else:
        return {"extraction_status": "unknown_error", "data": sanitize_text(str(extracted_data))}

def create_excel(all_results, output_structure, image_quality='high'):
    from openpyxl import Workbook
    from openpyxl.drawing import image as xl_image
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import tempfile
    excel_filename = f"json_extraction_{output_structure['topic']}_{output_structure['model_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_path = output_structure['base_output'] / excel_filename
    
    quality_settings = {
        'low': {'size': (150, 150), 'display': (150, 150), 'format': 'JPEG', 'quality': 70, 'row_height': 120, 'col_width': 20},
        'medium': {'size': (250, 250), 'display': (200, 200), 'format': 'JPEG', 'quality': 90, 'row_height': 160, 'col_width': 28},
        'high': {'size': (400, 400), 'display': (300, 300), 'format': 'PNG', 'quality': None, 'row_height': 240, 'col_width': 40}
    }
    settings = quality_settings[image_quality]
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON_Extraction"
        
        # Prepare data for Excel
        excel_data = []
        all_keys = set()
        
        for result in all_results:
            if 'error' not in result:
                row_data = {
                    'image_id': result['image_id'],
                    'image_name': result['image_name'],
                    'topic': output_structure['topic'],
                    'model_used': result['model_used'],
                    'provider': result['provider'],
                    'parsing_success': result.get('parsing_success', False),
                    'raw_json_response': result.get('raw_json_response', ''),
                    'file_size_mb': result['processing_metadata']['file_size_mb'],
                    'image_width': result['processing_metadata']['image_dimensions']['width'],
                    'image_height': result['processing_metadata']['image_dimensions']['height'],
                    'analysis_timestamp': result['analysis_timestamp'],
                    'image_path': result['image_path']
                }
                
                # Flatten the extracted data
                flattened_data = flatten_json_for_excel(result['extracted_data'])
                row_data.update(flattened_data)
                all_keys.update(flattened_data.keys())
                excel_data.append(row_data)
            else:
                error_row = {
                    'image_id': result.get('image_id', ''),
                    'image_name': result.get('image_name', ''),
                    'topic': output_structure['topic'],
                    'model_used': result.get('model_used', ''),
                    'provider': result.get('provider', ''),
                    'error': result['error'],
                    'analysis_timestamp': result.get('analysis_timestamp', ''),
                    'image_path': result.get('image_path', '')
                }
                excel_data.append(error_row)
        
        # Create headers - limit to reasonable number
        base_headers = ['Image', 'image_id', 'image_name', 'topic', 'model_used', 'provider', 
                       'parsing_success', 'raw_json_response', 'file_size_mb', 'image_width', 
                       'image_height', 'analysis_timestamp']
        data_headers = sorted(list(all_keys))
        
        # Limit total columns to prevent Excel issues
        max_cols = 50  # Safe limit for Excel
        if len(base_headers) + len(data_headers) > max_cols:
            data_headers = data_headers[:max_cols - len(base_headers)]
            print(f"⚠️ Limiting to {max_cols} columns for Excel compatibility")
        
        headers = base_headers + data_headers
        
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for idx, row_data in enumerate(excel_data):
            excel_row = [None]  # Image placeholder
            for header in headers[1:]:  # Skip 'Image' column
                excel_row.append(row_data.get(header, ''))
            ws.append(excel_row)
            
            # Add image if available
            img_path = row_data.get('image_path')
            if img_path and os.path.exists(img_path):
                try:
                    with Image.open(img_path) as pil_img:
                        if pil_img.mode in ('RGBA', 'P'):
                            pil_img = pil_img.convert('RGB')
                        pil_img.thumbnail(settings['size'], Image.Resampling.LANCZOS)
                        if settings['format'] == 'PNG':
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                                pil_img.save(tmp_file.name, 'PNG')
                                temp_img_path = tmp_file.name
                        else:
                            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
                                pil_img.save(tmp_file.name, 'JPEG', quality=settings['quality'])
                                temp_img_path = tmp_file.name
                                
                    img = xl_image.Image(temp_img_path)
                    img.width = settings['display'][0]
                    img.height = settings['display'][1]
                    cell_ref = f'A{idx + 2}'
                    ws.add_image(img, cell_ref)
                    ws.row_dimensions[idx + 2].height = settings['row_height']
                except Exception as e:
                    ws[f'A{idx + 2}'] = f"Error loading image"
        
        # Set column widths SAFELY - only for columns that exist
        try:
            ws.column_dimensions['A'].width = 40  # Image column
            for i in range(min(len(headers), 26)):  # Only first 26 columns (A-Z)
                col_letter = get_column_letter(i + 1)
                if i == 0:  # Image column
                    ws.column_dimensions[col_letter].width = 40
                elif headers[i] == 'raw_json_response':
                    ws.column_dimensions[col_letter].width = 50
                elif i <= 12:
                    ws.column_dimensions[col_letter].width = 15
                else:
                    ws.column_dimensions[col_letter].width = 20
        except Exception as e:
            print(f"⚠️ Warning: Could not set column widths: {e}")
        
        wb.save(excel_path)
        print(f"📊 Excel created: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        print(f"💾 Creating simple CSV instead...")
        
        # Create CSV as fallback
        try:
            import pandas as pd
            
            # Create simplified data for CSV
            csv_data = []
            for result in all_results:
                if 'error' not in result:
                    row = {
                        'image_id': result['image_id'],
                        'image_name': result['image_name'],
                        'model_used': result['model_used'],
                        'provider': result['provider'],
                        'parsing_success': result.get('parsing_success', False),
                        'analysis_timestamp': result['analysis_timestamp']
                    }
                    
                    # Add flattened data
                    flattened = flatten_json_for_excel(result['extracted_data'])
                    row.update(flattened)
                    csv_data.append(row)
            
            df = pd.DataFrame(csv_data)
            csv_path = output_structure['base_output'] / f"json_extraction_{output_structure['topic']}_{output_structure['model_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            df.to_csv(csv_path, index=False)
            print(f"📄 CSV created instead: {csv_path}")
            return csv_path
            
        except Exception as e2:
            print(f"❌ Error creating CSV file: {e2}")
            return None

def main():
    parser = argparse.ArgumentParser(description='JSON Advertisement Information Extractor - FIXED VERSION')
    parser.add_argument('--num_images', '-n', type=int, default=10, help='Number of images to process')
    parser.add_argument('--dataset_dir', '-d', type=str, default="pitt_ads/cars/images", help='Path to dataset directory')
    parser.add_argument('--model', '-m', type=str, default='gemini-2.0-flash-001', help='Model to use (auto-detects provider)')
    parser.add_argument('--provider', type=str, choices=['hf', 'gemini', 'openai'], help='Force provider: hf (HuggingFace), gemini, or openai')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    parser.add_argument('--image_quality', type=str, choices=['low', 'medium', 'high'], default='high', help='Image quality in Excel: low(150px,JPEG), medium(250px,JPEG), high(400px,PNG)')
    parser.add_argument('--device', type=str, default=None, help='Device to use for model and tensors (e.g., cuda:0, cpu). If not set, will use device_map="auto" for large models.')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    print(f"🔧 Available providers:")
    print(f"   Gemini: {'✅' if GEMINI_AVAILABLE else '❌'}")
    print(f"   OpenAI: {'✅' if OPENAI_AVAILABLE else '❌'}")
    print(f"   HuggingFace: {'✅' if HF_AVAILABLE else '❌'}")
    
    if HF_AVAILABLE and torch.cuda.is_available():
        print(f"🚀 GPU: {torch.cuda.get_device_name(0)}")
        if torch.cuda.device_count() > 1:
            print(f"🔥 Multiple GPUs: {torch.cuda.device_count()} devices")
    
    try:
        analyzer = get_analyzer(args.model, args.provider, args.device)
    except Exception as e:
        print(f"❌ Failed to initialize model {args.model}: {e}")
        return
    
    output_structure = create_output_structure(args.dataset_dir, args.model)
    
    ads_dir = Path(args.dataset_dir)
    if not ads_dir.exists():
        print(f"❌ Directory not found: {ads_dir}")
        return
    
    image_extensions = [".jpg", ".jpeg", ".png", ".bmp", ".tiff"]
    all_images = []
    for ext in image_extensions:
        all_images.extend(ads_dir.rglob(f"*{ext}"))
        all_images.extend(ads_dir.rglob(f"*{ext.upper()}"))
    
    selected_images = all_images[:args.num_images]
    
    print(f"📸 JSON Advertisement Information Extractor - FIXED VERSION")
    print(f"🤖 Model: {args.model}")
    print(f"📁 Dataset: {ads_dir}")
    print(f"🖼️ Images: {len(selected_images)}")
    print(f"📝 Prompt: JSON structured extraction")
    print(f"🗂️ Output will be saved to: {output_structure['output_dir']}")
    print("=" * 70)
    
    all_results = []
    start_time = datetime.now()
    
    for i, img_path in enumerate(selected_images, 1):
        print(f"\n[{i}/{len(selected_images)}] 🔍 Processing: {img_path.name}")
        result = analyzer.analyze_image_json_extraction(img_path)
        all_results.append(result)
        print(f"   ✅ Completed JSON extraction for {img_path.name}")
        print("-" * 50)
    
    print(f"\n💾 Saving individual JSON files...")
    saved_files = save_results_to_folder(all_results, output_structure)
    print(f"   {len(saved_files)} files saved")
    
    print(f"\n📊 Creating Excel file...")
    excel_path = create_excel(all_results, output_structure, args.image_quality)
    
    duration = datetime.now() - start_time
    print(f"\n🎉 Analysis Complete!")
    print(f"⏱️ Duration: {duration}")
    print(f"🖼️ Images processed: {len(selected_images)}")
    print(f"🔧 API calls: {analyzer.api_calls_count}")
    print(f"📊 Excel file: {excel_path}")

if __name__ == "__main__":
    main()