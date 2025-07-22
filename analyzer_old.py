#!/usr/bin/env python3
"""
Universal Excel Multi-Prompt Analyzer
Supports: Gemini API (new google-genai SDK), OpenAI API, and Hugging Face local models
- Auto-detects provider based on model name
- Same output format for all providers
- Single Excel sheet with all prompts as columns
"""

import os
import json
import argparse
import shutil
from pathlib import Path
from PIL import Image
import time
import logging
from datetime import datetime
import pandas as pd
import base64
import io

# Try importing all providers
try:
    from google import genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import torch
    from transformers import (
        LlavaNextProcessor, LlavaNextForConditionalGeneration,
        LlavaForConditionalGeneration, LlavaProcessor,
        AutoProcessor, AutoModelForCausalLM
    )
    HF_AVAILABLE = True
except ImportError:
    HF_AVAILABLE = False

# Simple .env loader
def load_env_file():
    try:
        with open('.env', 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    value = value.strip().strip('"').strip("'")
                    os.environ[key.strip()] = value
    except FileNotFoundError:
        pass

load_env_file()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BaseVisionAnalyzer:
    """Base class for all vision analyzers"""
    
    def __init__(self, model_name):
        self.model_name = model_name
        self.api_calls_count = 0
        
        # Define all three prompts
        self.prompts = {
            'basic': """Describe what you see in this advertisement image, focusing on:
- Main visual elements
- Product being advertised
- Key text or messages
- Overall visual style""",
            
            'detailed': """Describe this advertisement image in detail, including:
- All visible objects and elements
- Text and branding elements
- Colors and visual composition
- The product or service being advertised
- Target audience and marketing approach""",
            
            'comprehensive': """Provide a comprehensive description of this advertisement image focusing on:

1. VISUAL ELEMENTS:
- Objects, people, products visible
- Colors, lighting, composition style
- Text, logos, slogans present
- Background and setting

2. ADVERTISING CONTEXT:
- What product/service is being advertised
- Brand indicators and visual identity
- Target audience signals
- Overall mood and aesthetic tone

3. MARKETING ELEMENTS:
- Visual storytelling approach
- Emotional appeals used
- Product positioning strategy
- Call-to-action elements (if any)

Provide a detailed, professional description suitable for marketing analysis."""
        }
    
    def analyze_image_all_prompts(self, image_path):
        """Run all three prompts on the same image - to be implemented by subclasses"""
        raise NotImplementedError
    
    def _get_image_dimensions(self, image_path):
        """Get image dimensions"""
        try:
            with Image.open(image_path) as img:
                return {"width": img.width, "height": img.height}
        except:
            return {"width": 0, "height": 0}

class GeminiAnalyzer(BaseVisionAnalyzer):
    """Gemini API analyzer using new Google GenAI SDK"""
    
    def __init__(self, model_name="gemini-2.0-flash-001"):
        super().__init__(model_name)
        self._initialize_model()
    
    def _initialize_model(self):
        if not GEMINI_AVAILABLE:
            raise ImportError("google-genai not installed. Run: pip install google-genai")
        
        try:
            api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
            if not api_key:
                raise ValueError("GEMINI_API_KEY or GOOGLE_API_KEY not found in environment")
            
            # Create client with new SDK
            self.client = genai.Client(api_key=api_key)
            logger.info(f"✅ Gemini model initialized: {self.model_name}")
        except Exception as e:
            logger.error(f"❌ Failed to initialize Gemini: {e}")
            raise
    
    def analyze_image_all_prompts(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        
        logger.info(f"Analyzing: {image_name} with ALL prompts (Gemini)")
        
        results = {}
        
        for prompt_type, prompt_text in self.prompts.items():
            try:
                print(f"   🔄 Running {prompt_type} prompt...")
                
                # Read and encode image for new SDK
                with open(image_path, 'rb') as f:
                    image_data = base64.b64encode(f.read()).decode('utf-8')
                
                # Determine image format
                image_format = image_path.suffix.lower().replace('.', '')
                if image_format == 'jpg':
                    image_format = 'jpeg'
                
                # Create content for new SDK
                response = self.client.models.generate_content(
                    model=self.model_name,
                    contents=[
                        prompt_text,
                        {"inline_data": {"mime_type": f"image/{image_format}", "data": image_data}}
                    ]
                )
                
                description = response.text.strip()
                self.api_calls_count += 1
                
                result = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'gemini',
                    'prompt_type': prompt_type,
                    'description': description,
                    'description_length': len(description),
                    'analysis_timestamp': datetime.now().isoformat(),
                    'processing_metadata': {
                        'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                        'image_dimensions': self._get_image_dimensions(image_path)
                    }
                }
                
                results[prompt_type] = result
                print(f"   ✅ {prompt_type}: {len(description)} chars")
                time.sleep(1)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error with {prompt_type} prompt: {e}")
                results[prompt_type] = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'gemini',
                    'prompt_type': prompt_type,
                    'error': str(e),
                    'analysis_timestamp': datetime.now().isoformat()
                }
        
        return results

class OpenAIAnalyzer(BaseVisionAnalyzer):
    """OpenAI API analyzer"""
    
    def __init__(self, model_name="gpt-4o"):
        super().__init__(model_name)
        self._initialize_model()
    
    def _initialize_model(self):
        if not OPENAI_AVAILABLE:
            raise ImportError("openai not installed. Run: pip install openai")
        
        try:
            api_key = os.getenv("OPENAI_API_KEY")
            if not api_key:
                raise ValueError("OPENAI_API_KEY not found in environment")
            self.client = openai.OpenAI(api_key=api_key)
            logger.info(f"✅ OpenAI model initialized: {self.model_name}")
        except Exception as e:
            logger.error(f"❌ Failed to initialize OpenAI: {e}")
            raise
    
    def _encode_image_base64(self, image_path):
        """Encode image to base64 for OpenAI"""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    
    def analyze_image_all_prompts(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        
        logger.info(f"Analyzing: {image_name} with ALL prompts (OpenAI)")
        
        results = {}
        
        # Encode image once
        base64_image = self._encode_image_base64(image_path)
        
        for prompt_type, prompt_text in self.prompts.items():
            try:
                print(f"   🔄 Running {prompt_type} prompt...")
                
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt_text},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/jpeg;base64,{base64_image}"
                                    }
                                }
                            ]
                        }
                    ],
                    max_tokens=1000
                )
                
                description = response.choices[0].message.content.strip()
                self.api_calls_count += 1
                
                result = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'openai',
                    'prompt_type': prompt_type,
                    'description': description,
                    'description_length': len(description),
                    'analysis_timestamp': datetime.now().isoformat(),
                    'processing_metadata': {
                        'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                        'image_dimensions': self._get_image_dimensions(image_path)
                    }
                }
                
                results[prompt_type] = result
                print(f"   ✅ {prompt_type}: {len(description)} chars")
                time.sleep(1)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error with {prompt_type} prompt: {e}")
                results[prompt_type] = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'openai',
                    'prompt_type': prompt_type,
                    'error': str(e),
                    'analysis_timestamp': datetime.now().isoformat()
                }
        
        return results

class HuggingFaceAnalyzer(BaseVisionAnalyzer):
    """Hugging Face local model analyzer"""
    
    def __init__(self, model_name="llava-hf/llava-1.5-13b-hf"):
        super().__init__(model_name)
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.model = None
        self.processor = None
        self._initialize_model()
    
    def _initialize_model(self):
        if not HF_AVAILABLE:
            raise ImportError("transformers/torch not installed. Run: pip install transformers torch")
        
        try:
            print(f"🔄 Loading model: {self.model_name}")
            print(f"📍 Device: {self.device}")
            
            # Check if model is cached
            cache_dir = os.path.expanduser("~/.cache/huggingface/hub")
            model_cache_name = f"models--{self.model_name.replace('/', '--')}"
            model_cache_path = os.path.join(cache_dir, model_cache_name)
            
            if os.path.exists(model_cache_path):
                print(f"💾 Found cached model at: {model_cache_path}")
                print(f"📦 Cache size: {self._get_folder_size(model_cache_path)}")
            else:
                print(f"⬇️ Model not cached - will download to: {model_cache_path}")
            
            if "llava-v1.6" in self.model_name or "llava-next" in self.model_name:
                self.processor = LlavaNextProcessor.from_pretrained(self.model_name, use_fast=True)
                self.model = LlavaNextForConditionalGeneration.from_pretrained(
                    self.model_name,
                    torch_dtype=torch.float16,
                    device_map="auto",
                    low_cpu_mem_usage=True
                )
            elif "llava" in self.model_name:
                self.processor = LlavaProcessor.from_pretrained(self.model_name, use_fast=True)
                self.model = LlavaForConditionalGeneration.from_pretrained(
                    self.model_name,
                    torch_dtype=torch.float16,
                    device_map="auto",
                    low_cpu_mem_usage=True
                )
            else:
                self.processor = AutoProcessor.from_pretrained(self.model_name, use_fast=True)
                self.model = AutoModelForCausalLM.from_pretrained(
                    self.model_name,
                    torch_dtype=torch.float16,
                    device_map="auto",
                    low_cpu_mem_usage=True
                )
            
            logger.info(f"✅ Hugging Face model loaded successfully")
            
        except Exception as e:
            logger.error(f"❌ Failed to load Hugging Face model: {e}")
            raise
    
    def analyze_image_all_prompts(self, image_path):
        image_id = Path(image_path).stem
        image_name = Path(image_path).name
        
        logger.info(f"Analyzing: {image_name} with ALL prompts (HuggingFace)")
        
        results = {}
        image = Image.open(image_path).convert('RGB')
        
        for prompt_type, prompt_text in self.prompts.items():
            try:
                print(f"   🔄 Running {prompt_type} prompt...")
                
                description = self._generate_description(image, prompt_text)
                
                result = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'huggingface',
                    'prompt_type': prompt_type,
                    'description': description,
                    'description_length': len(description),
                    'analysis_timestamp': datetime.now().isoformat(),
                    'processing_metadata': {
                        'file_size_mb': round(os.path.getsize(image_path) / (1024 * 1024), 2),
                        'image_dimensions': self._get_image_dimensions(image_path)
                    }
                }
                
                results[prompt_type] = result
                print(f"   ✅ {prompt_type}: {len(description)} chars")
                time.sleep(0.5)
                
            except Exception as e:
                logger.error(f"Error with {prompt_type} prompt: {e}")
                results[prompt_type] = {
                    'image_path': str(image_path),
                    'image_name': image_name,
                    'image_id': image_id,
                    'model_used': self.model_name,
                    'provider': 'huggingface',
                    'prompt_type': prompt_type,
                    'error': str(e),
                    'analysis_timestamp': datetime.now().isoformat()
                }
        
        return results
    
    def _generate_description(self, image, prompt):
        try:
            if "llava" in self.model_name:
                full_prompt = f"USER: <image>\n{prompt}\nASSISTANT:"
                inputs = self.processor(text=full_prompt, images=image, return_tensors="pt")
                
                for key in inputs:
                    if torch.is_tensor(inputs[key]):
                        inputs[key] = inputs[key].to(self.device)
                
                with torch.no_grad():
                    outputs = self.model.generate(
                        **inputs,
                        max_new_tokens=512,
                        do_sample=True,
                        temperature=0.7,
                        pad_token_id=self.processor.tokenizer.eos_token_id
                    )
                
                response = self.processor.decode(outputs[0], skip_special_tokens=True)
                
                if "ASSISTANT:" in response:
                    description = response.split("ASSISTANT:")[-1].strip()
                else:
                    description = response.strip()
                    
            else:
                inputs = self.processor(images=image, text=prompt, return_tensors="pt")
                
                for key in inputs:
                    if torch.is_tensor(inputs[key]):
                        inputs[key] = inputs[key].to(self.device)
                
                with torch.no_grad():
                    outputs = self.model.generate(**inputs, max_new_tokens=512)
                
                description = self.processor.decode(outputs[0], skip_special_tokens=True)
            
            return description
            
        except Exception as e:
            return f"Error: {str(e)}"
    
    def _get_folder_size(self, folder_path):
        """Get human readable folder size"""
        try:
            total_size = 0
            for dirpath, dirnames, filenames in os.walk(folder_path):
                for f in filenames:
                    fp = os.path.join(dirpath, f)
                    if os.path.exists(fp):
                        total_size += os.path.getsize(fp)
            
            # Convert to human readable
            for unit in ['B', 'KB', 'MB', 'GB']:
                if total_size < 1024.0:
                    return f"{total_size:.1f} {unit}"
                total_size /= 1024.0
            return f"{total_size:.1f} TB"
        except:
            return "Unknown size"

def get_analyzer(model_name, provider=None):
    """Factory function to get the right analyzer based on model name"""
    
    # Check for Hugging Face models
    if provider == 'hf':
        if not HF_AVAILABLE:
            raise ImportError("Hugging Face model requested but transformers/torch not installed")
        return HuggingFaceAnalyzer(model_name)
    
    # Check for OpenAI models
    elif provider == 'openai':
        if not OPENAI_AVAILABLE:
            raise ImportError("OpenAI model requested but openai library not installed")
        return OpenAIAnalyzer(model_name)
    
    # Check for Gemini models (includes new model names)
    elif provider == 'gemini':
        if not GEMINI_AVAILABLE:
            raise ImportError("Gemini model requested but google-genai not installed")
        return GeminiAnalyzer(model_name)
    
    # Default to Gemini for unknown models
    else:
        if not GEMINI_AVAILABLE:
            raise ImportError("Default Gemini model requested but google-genai not installed")
        return GeminiAnalyzer(model_name)

# Keep existing helper functions
def create_output_structure_with_prompts(dataset_dir, model_name, topic=None):
    if topic is None:
        dataset_path = Path(dataset_dir)
        if "cars" in str(dataset_path):
            topic = "cars"
        elif "electronics" in str(dataset_path):
            topic = "electronics"
        elif "fashion" in str(dataset_path):
            topic = "fashion"
        else:
            parts = dataset_path.parts
            if len(parts) >= 2:
                topic = parts[-2]
            else:
                topic = "general"
    
    clean_model_name = model_name.replace("/", "_").replace(":", "_").replace("-", "_")
    
    output_dirs = {}
    prompt_types = ['basic', 'detailed', 'comprehensive']
    
    for prompt_type in prompt_types:
        output_dir = Path("output") / prompt_type / clean_model_name / topic
        output_dir.mkdir(parents=True, exist_ok=True)
        output_dirs[prompt_type] = output_dir
    
    return {
        'output_dirs': output_dirs,
        'topic': topic,
        'model_name': clean_model_name,
        'base_output': Path("output")
    }

def save_results_to_folders(all_results, output_structure):
    saved_files = {'basic': [], 'detailed': [], 'comprehensive': []}
    
    for image_results in all_results:
        for prompt_type, result in image_results.items():
            if 'error' not in result:
                output_dir = output_structure['output_dirs'][prompt_type]
                filename = f"{result['image_id']}.json"
                file_path = output_dir / filename
                
                with open(file_path, 'w') as f:
                    json.dump(result, f, indent=2)
                
                saved_files[prompt_type].append(file_path)
    
    return saved_files

def create_comprehensive_excel(all_results, output_structure, image_quality='high'):
    from openpyxl import Workbook
    from openpyxl.drawing import image as xl_image
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment
    import tempfile
    
    all_data = []
    image_paths = []  # Keep track of image paths for embedding
    
    for image_results in all_results:
        row_data = {}
        current_image_path = None
        
        first_result = None
        for result in image_results.values():
            if 'error' not in result:
                first_result = result
                break
        
        if first_result:
            current_image_path = first_result['image_path']
            row_data['image_id'] = first_result['image_id']
            row_data['image_name'] = first_result['image_name']
            row_data['model_used'] = first_result['model_used']
            row_data['provider'] = first_result.get('provider', 'unknown')
            row_data['analysis_timestamp'] = first_result['analysis_timestamp']
            row_data['file_size_mb'] = first_result['processing_metadata']['file_size_mb']
            row_data['image_width'] = first_result['processing_metadata']['image_dimensions']['width']
            row_data['image_height'] = first_result['processing_metadata']['image_dimensions']['height']
            row_data['topic'] = output_structure['topic']
        
        for prompt_type in ['basic', 'detailed', 'comprehensive']:
            if prompt_type in image_results and 'error' not in image_results[prompt_type]:
                result = image_results[prompt_type]
                row_data[f'{prompt_type}_description'] = result['description']
                row_data[f'{prompt_type}_length'] = result['description_length']
            else:
                row_data[f'{prompt_type}_description'] = 'ERROR' if prompt_type in image_results else 'MISSING'
                row_data[f'{prompt_type}_length'] = 0
        
        all_data.append(row_data)
        image_paths.append(current_image_path)
    
    excel_filename = f"analysis_results_{output_structure['topic']}_{output_structure['model_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_path = output_structure['base_output'] / excel_filename
    
    # Configure image settings based on quality
    quality_settings = {
        'low': {'size': (150, 150), 'display': (150, 150), 'format': 'JPEG', 'quality': 70, 'row_height': 120, 'col_width': 20},
        'medium': {'size': (250, 250), 'display': (200, 200), 'format': 'JPEG', 'quality': 90, 'row_height': 160, 'col_width': 28},
        'high': {'size': (400, 400), 'display': (300, 300), 'format': 'PNG', 'quality': None, 'row_height': 240, 'col_width': 40}
    }
    
    settings = quality_settings[image_quality]
    
    if all_data:
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Column order - put image column first
        column_order = [
            'image_id', 'image_name', 'topic', 'model_used', 'provider',
            'basic_description', 'basic_length',
            'detailed_description', 'detailed_length', 
            'comprehensive_description', 'comprehensive_length',
            'file_size_mb', 'image_width', 'image_height',
            'analysis_timestamp'
        ]
        
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        # Create workbook manually to add images
        wb = Workbook()
        ws = wb.active
        ws.title = "All_Results"
        
        # Add an "Image" column at the beginning
        headers = ['Image'] + list(df.columns)
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows (skip image column for now)
        for idx, row in df.iterrows():
            ws.append([None] + list(row))  # None for image column
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Image column - wider for larger images
        ws.column_dimensions['B'].width = 15  # image_id
        ws.column_dimensions['C'].width = 20  # image_name
        ws.column_dimensions['D'].width = 12  # topic
        ws.column_dimensions['E'].width = 20  # model_used
        ws.column_dimensions['F'].width = 12  # provider
        
        # Description columns - make them wider but limit height
        for col in ['G', 'I', 'K']:  # basic, detailed, comprehensive description columns
            ws.column_dimensions[col].width = 50
        
        print(f"📊 Adding images to Excel...")
        
        # Keep track of temp files to clean up later
        temp_files = []
        
        # Add images to each row
        for idx, img_path in enumerate(image_paths):
            if img_path and os.path.exists(img_path):
                try:
                    # Create a resized copy of the image
                    with Image.open(img_path) as pil_img:
                        # Convert to RGB if needed (for PNG with transparency)
                        if pil_img.mode in ('RGBA', 'P'):
                            pil_img = pil_img.convert('RGB')
                        
                        # Resize based on quality setting
                        pil_img.thumbnail(settings['size'], Image.Resampling.LANCZOS)
                        
                        # Save with appropriate format and quality
                        if settings['format'] == 'PNG':
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                                pil_img.save(tmp_file.name, 'PNG')
                                temp_img_path = tmp_file.name
                                temp_files.append(temp_img_path)
                        else:  # JPEG
                            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
                                pil_img.save(tmp_file.name, 'JPEG', quality=settings['quality'])
                                temp_img_path = tmp_file.name
                                temp_files.append(temp_img_path)
                    
                    # Add image to Excel with quality-based display size
                    img = xl_image.Image(temp_img_path)
                    img.width = settings['display'][0]
                    img.height = settings['display'][1]
                    
                    # Position in the image column (A) for this row
                    cell_ref = f'A{idx + 2}'  # +2 because row 1 is headers, and idx starts at 0
                    ws.add_image(img, cell_ref)
                    
                    # Set row height based on quality setting
                    ws.row_dimensions[idx + 2].height = settings['row_height']
                    
                except Exception as e:
                    print(f"   ⚠️ Failed to add image {img_path}: {e}")
                    # Add text instead
                    ws[f'A{idx + 2}'] = f"Error loading image"
        
        # Save workbook (this is when Excel actually reads the image files)
        wb.save(excel_path)
        
        # Now clean up temp files
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass  # Ignore cleanup errors
        
        print(f"📊 Excel created with {len(df)} rows, {len(df.columns)} columns, and high-quality embedded images")
        print(f"🖼️ Image settings: {settings['size'][0]}x{settings['size'][1]}px max, display: {settings['display'][0]}x{settings['display'][1]}px, format: {settings['format']}")
        if settings['quality']:
            print(f"🎨 JPEG quality: {settings['quality']}%")
    
    return excel_path

def show_cache_info():
    """Show Hugging Face cache information"""
    cache_dir = os.path.expanduser("~/.cache/huggingface/hub")
    
    if not os.path.exists(cache_dir):
        print("📦 No Hugging Face cache found")
        return
    
    print(f"📦 Cache location: {cache_dir}")
    
    # Get total cache size and list models
    total_size = 0
    model_info = []
    
    try:
        for item in os.listdir(cache_dir):
            if item.startswith("models--"):
                # Convert cache name back to model name
                model_name = item.replace("models--", "").replace("--", "/")
                
                item_path = os.path.join(cache_dir, item)
                model_size = 0
                
                for dirpath, dirnames, filenames in os.walk(item_path):
                    for f in filenames:
                        fp = os.path.join(dirpath, f)
                        if os.path.exists(fp):
                            file_size = os.path.getsize(fp)
                            model_size += file_size
                            total_size += file_size
                
                # Convert to human readable
                size_str = model_size
                for unit in ['B', 'KB', 'MB', 'GB']:
                    if size_str < 1024.0:
                        size_str = f"{size_str:.1f} {unit}"
                        break
                    size_str /= 1024.0
                else:
                    size_str = f"{size_str:.1f} TB"
                
                model_info.append((model_name, size_str))
        
        # Convert total to human readable
        total_str = total_size
        for unit in ['B', 'KB', 'MB', 'GB']:
            if total_str < 1024.0:
                total_str = f"{total_str:.1f} {unit}"
                break
            total_str /= 1024.0
        else:
            total_str = f"{total_str:.1f} TB"
        
        print(f"📊 Total cache size: {total_str}")
        print(f"🤖 Cached models: {len(model_info)}")
        
        if model_info:
            print("\n📋 Cached models:")
            for model_name, size in sorted(model_info):
                print(f"   {model_name}: {size}")
        
    except Exception as e:
        print(f"❌ Error reading cache: {e}")

def main():
    parser = argparse.ArgumentParser(description='Universal Vision Multi-Prompt Analyzer')
    parser.add_argument('--num_images', '-n', type=int, default=10,
                       help='Number of images to process')
    parser.add_argument('--dataset_dir', '-d', type=str,
                       default="pitt_ads/cars/images",
                       help='Path to dataset directory')
    parser.add_argument('--model', '-m', type=str, default='gemini-2.0-flash-001',
                       help='Model to use (auto-detects provider)')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Verbose output')
    parser.add_argument('--image_quality', type=str, choices=['low', 'medium', 'high'], default='high',
                       help='Image quality in Excel: low(150px,JPEG), medium(250px,JPEG), high(400px,PNG)')
    parser.add_argument('--show_cache', action='store_true',
                       help='Show HuggingFace cache info and exit')
    parser.add_argument('--clear_cache', type=str, metavar='MODEL_NAME',
                       help='Clear cache for specific model and exit')
    parser.add_argument('--provider', type=str, choices=['hf', 'gemini', 'openai'], help='Force provider: hf (HuggingFace), gemini, or openai')
    
    args = parser.parse_args()
    
    # Handle cache management options
    if args.show_cache:
        show_cache_info()
        return
    
    if args.clear_cache:
        cache_dir = os.path.expanduser("~/.cache/huggingface/hub")
        model_cache_name = f"models--{args.clear_cache.replace('/', '--')}"
        model_cache_path = os.path.join(cache_dir, model_cache_name)
        
        if os.path.exists(model_cache_path):
            import shutil
            shutil.rmtree(model_cache_path)
            print(f"✅ Cleared cache for: {args.clear_cache}")
        else:
            print(f"❌ No cache found for: {args.clear_cache}")
        return
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Show available providers
    print(f"🔧 Available providers:")
    print(f"   Gemini: {'✅' if GEMINI_AVAILABLE else '❌'}")
    print(f"   OpenAI: {'✅' if OPENAI_AVAILABLE else '❌'}")
    print(f"   HuggingFace: {'✅' if HF_AVAILABLE else '❌'}")
    
    if HF_AVAILABLE and torch.cuda.is_available():
        print(f"🚀 GPU: {torch.cuda.get_device_name(0)}")
        if torch.cuda.device_count() > 1:
            print(f"🔥 Multiple GPUs: {torch.cuda.device_count()} devices")
    
    # Show cache info for HuggingFace models
    if any(hf_indicator in args.model.lower() for hf_indicator in 
           ['llava', 'qwen', 'git-', 'blip', 'instructblip', 'fuyu']):
        print("\n" + "="*50)
        show_cache_info()
        print("="*50)
    
    # Initialize analyzer
    try:
        analyzer = get_analyzer(args.model, args.provider)
    except Exception as e:
        print(f"❌ Failed to initialize model {args.model}: {e}")
        return
    
    # Create output structure
    output_structure = create_output_structure_with_prompts(args.dataset_dir, args.model)
    
    # Find images
    ads_dir = Path(args.dataset_dir)
    if not ads_dir.exists():
        print(f"❌ Directory not found: {ads_dir}")
        return
    
    image_extensions = [".jpg", ".jpeg", ".png", ".bmp", ".tiff"]
    all_images = []
    for ext in image_extensions:
        all_images.extend(ads_dir.rglob(f"*{ext}"))
        all_images.extend(ads_dir.rglob(f"*{ext.upper()}"))
    
    selected_images = all_images[:args.num_images]
    
    print(f"📸 Universal Multi-Prompt Analyzer")
    print(f"🤖 Model: {args.model}")
    print(f"📁 Dataset: {ads_dir}")
    print(f"🖼️ Images: {len(selected_images)}")
    print(f"📝 Prompts: basic, detailed, comprehensive")
    print(f"🗂️ Output will be saved to:")
    for prompt_type, folder in output_structure['output_dirs'].items():
        print(f"   {prompt_type}: {folder}")
    print("=" * 70)
    
    all_results = []
    start_time = datetime.now()
    
    for i, img_path in enumerate(selected_images, 1):
        print(f"\n[{i}/{len(selected_images)}] 🔍 Processing: {img_path.name}")
        
        image_results = analyzer.analyze_image_all_prompts(img_path)
        all_results.append(image_results)
        
        print(f"   ✅ Completed all 3 prompts for {img_path.name}")
        print("-" * 50)
    
    # Save results
    print(f"\n💾 Saving individual JSON files...")
    saved_files = save_results_to_folders(all_results, output_structure)
    
    for prompt_type, files in saved_files.items():
        print(f"   {prompt_type}: {len(files)} files saved")
    
    print(f"\n📊 Creating comprehensive Excel file...")
    excel_path = create_comprehensive_excel(all_results, output_structure, args.image_quality)
    
    duration = datetime.now() - start_time
    print(f"\n🎉 Analysis Complete!")
    print(f"⏱️ Duration: {duration}")
    print(f"🖼️ Images processed: {len(selected_images)}")
    print(f"🔧 API calls: {analyzer.api_calls_count}")
    print(f"📊 Excel file: {excel_path}")

if __name__ == "__main__":
    main()